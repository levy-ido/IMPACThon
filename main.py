import openpyxl
from google.cloud import translate_v2 as translate


def translate_descriptions():
    translate_client = translate.Client.from_service_account_json('singular-carver-388607-9151ff88d8f7.json')
    workbook = openpyxl.load_workbook('guidestar_spreadsheet.xlsx')
    sheet = workbook.active
    max_row = sheet.max_row
    for row in range(1, max_row + 1):
        hebrew_text = sheet.cell(row=row, column=6).value
        english_text = "Error"
        try:
            translation = translate_client.translate(hebrew_text,target_language='en')
            english_text = translation['translatedText']
        except Exception as e:
            print(f"Translation error in row {row}")
        sheet.cell(row=row, column=6).value = english_text

    workbook.save('translated_spreadsheet.xlsx')


if __name__ == '__main__':
    translate_descriptions()

